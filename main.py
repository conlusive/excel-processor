import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def process_excel(input_file, output_file):
    df = pd.read_excel(input_file, engine="openpyxl")

    numeric_cols = df.select_dtypes(include=['number']).columns
    df['Average'] = df[numeric_cols].mean(axis=1)

    if len(numeric_cols) > 0:
        df_sorted = df.sort_values(by=numeric_cols[0], ascending=False)
    else:
        df_sorted = df

    df_sorted.to_excel(output_file, index=False, engine="openpyxl")

    wb = load_workbook(output_file)
    ws = wb.active

    chart = BarChart()
    chart.title = "Scores"
    chart.x_axis.title = "Name"
    chart.y_axis.title = "Score"

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E2")
    wb.save(output_file)

    print(f"\n The result is saved in {output_file} with a graph and a column")

if __name__ == "__main__":
    process_excel("input.xlsx", "output.xlsx")

